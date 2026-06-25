import os
import sys
import subprocess

def install_and_import(package):
    try:
        __import__(package)
    except ImportError:
        print(f"Installing {package}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Ensure openpyxl is installed
install_and_import('openpyxl')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_user_stories_excel():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # -------------------------------------------------------------
    # 1. Hotel Management System Sheet
    # -------------------------------------------------------------
    ws_hotel = wb.create_sheet(title="Hotel Management System")
    ws_hotel.views.sheetView[0].showGridLines = True
    
    # Title Row
    ws_hotel.merge_cells('A1:D1')
    title_cell = ws_hotel['A1']
    title_cell.value = "HOTEL MANAGEMENT SYSTEM — USER STORIES & ACCEPTANCE CRITERIA"
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_hotel.row_dimensions[1].height = 40
    
    # Headers
    headers = ["S.N.", "Feature Name", "User Story", "Acceptance Criteria"]
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_hotel.cell(row=2, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    ws_hotel.row_dimensions[2].height = 25
    
    hotel_stories = [
        (
            2,
            "Secure Registration & Login",
            "As a user, I want to register and log in with role selection so that I can access role-based features securely.",
            "• The system allows users to select a role (Admin, Staff, Customer) during registration.\n• User passwords are stored securely.\n• The system validates required fields.\n• Users only get access features permitted to their assigned role."
        ),
        (
            3,
            "Dashboard Overview",
            "As a manager, I want to view a dashboard overview so that I can quickly monitor hotel operations such as bookings, room status, and revenue.",
            "• The dashboard displays summarized data such as total bookings, available rooms, and revenue.\n• The dashboard reflects all the updates.\n• The system displays a different dashboard based on user roles.\n• Dashboard loads quickly and is accessible immediately after login."
        ),
        (
            4,
            "Booking Management",
            "As a manager, I want to manage room bookings so that I can reserve, update, or cancel rooms efficiently.",
            "• The manager can view all room bookings (confirmed, pending, and cancelled) in a centralized dashboard.\n• The system allows the manager to add, update, or remove room availability.\n• Error messages will be displayed for invalid booking operations.\n• The manager can cancel bookings due to operational reasons."
        ),
        (
            5,
            "System Setting and control",
            "As a manager, I want to manage system settings so that I can update hotel information, currency and pricing all together.",
            "• The manager can update hotel details such as name, address, email and contacts.\n• The manager can select and update the currency symbol used across the system for pricing and invoices.\n• The manager can update room service pricing, and changes are immediately reflected in billing and invoices.\n• The manager can set and modify tax rates and ensure they are applied to billing operations."
        ),
        (
            6,
            "Room Management",
            "As a manager, I want to add, update, and delete room details so that room inventory is maintained.",
            "• Admin can create rooms with numbers, type, price.\n• Admin can update room details.\n• Admin can delete rooms.\n• Changes can reflect in real-time."
        ),
        (
            7,
            "Search guest(by id)",
            "As a frontdesk staff, I want to search for guest details so that I can quickly access their booking information.",
            "• Frontdesk can search by name or id.\n• System returns matching results.\n• Guest details are displayed clearly.\n• Search works quickly."
        ),
        (
            8,
            "Room Status Checking",
            "As frontdesk staff, I want to check room status so that I know which rooms are available, occupied, or under maintenance.",
            "• The system shows the room availability status.\n• Rooms will be categorized by available, occupied or under maintenance.\n• Status updates in real-time in the system.\n• Staff can view the room status dashboard."
        ),
        (
            9,
            "Staff Management",
            "As a manager, I want to manage staff accounts so that roles and responsibilities are controlled.",
            "• The manager can create new staff accounts with required details.\n• The manager can assign and update roles for each staff member.\n• The manager can view a list of all staff along with their roles.\n• The manager can deactivate or remove staff accounts when required."
        ),
        (
            10,
            "Manage Customer Accounts",
            "As a hotel manager, I want to manage customer accounts so that I can view, update, and control guest profiles and their associated information efficiently.",
            "• The manager can view guest accounts by name or ID.\n• The manager can update customer profile details such as contact information and preferences.\n• The manager can deactivate or delete a customer account.\n• All changes are saved in real-time and access is restricted to authorized roles only."
        ),
        (
            11,
            "View Reports",
            "As a hotel manager, I want to view operational reports so that I can track revenue, occupancy, and bookings to make informed decisions.",
            "• The manager can access categorized reports (Revenue, Occupancy, Bookings) from the dashboard.\n• Reports can be filtered by date range (daily, weekly, monthly, or custom).\n• Each report displays accurate, up-to-date data in a clear and readable format.\n• The manager can export or print reports in PDF format."
        ),
        (
            12,
            "Discount & Offer Management",
            "As a hotel manager, I want to create and manage discounts and offers so that I can apply promotional pricing and attract more guests.",
            "• The manager can create a discount with a name, type (percentage or fixed), value, and validity period.\n• The manager can assign a unique promo code applicable during booking or checkout.\n• The system automatically deducts the correct amount when a valid code is applied.\n• Expired or invalid codes are rejected with a clear error message."
        ),
        (
            13,
            "Billing & Invoicing",
            "As a hotel manager, I want to generate and manage invoices so that I can accurately bill guests for all services and maintain clear payment records.",
            "• The system automatically generates an invoice at checkout with a full breakdown of all charges.\n• Taxes and discounts are calculated and displayed as separate line items on the invoice.\n• The manager can add, edit, or remove charges before finalizing the invoice.\n• Finalized invoices can be downloaded as PDF and are stored in invoice history."
        ),
        (
            14,
            "Room Service Request",
            "As a guest, I want to request room service through the system so that I can request for cleaning the room when I want.",
            "• The guest can submit a room service request for cleaning through the system.\n• The system records the request with room number, date, and time.\n• The guest receives confirmation after successfully submitting the request.\n• The request is visible to staff for processing and status updates (e.g., pending, completed)."
        ),
        (
            15,
            "Room Service & Housekeeping Tracker",
            "As a hotel manager, I want to track room service and housekeeping requests so that I can manage tasks efficiently and avoid delays.",
            "• The staff dashboard displays all pending service or housekeeping requests.\n• Requests show guest details, room number, and task type.\n• Staff can update status (e.g., accepted, in-progress, completed).\n• System logs timestamps for request creation and completion."
        ),
        (
            16,
            "Order Requests",
            "As a guest, I want to place different types of orders (food, drinks, amenities) so that I can enjoy hotel services conveniently.",
            "• Guests can browse categorized items (food, drinks, amenities).\n• Orders can be customized (e.g., add-ons, portion size).\n• The system calculates total cost before confirmation.\n• Guests receive digital confirmation of the order."
        ),
        (
            17,
            "View Expenses",
            "As a guest, I want to view my expenses during my stay so that I can monitor my spending and prepare for checkout.",
            "• Guests can view a detailed list of all charges (room, food, services).\n• Expenses are updated in real-time after each transaction.\n• The system shows breakdown by category (room, dining, housekeeping, extras).\n• Guests can download or request a digital copy of expenses."
        ),
        (
            18,
            "Staff Attendance & Shift Monitoring",
            "As a manager, I want to track staff attendance and shifts so that I can manage employee working hours efficiently.",
            "• Staff can easily check in and check out without confusion.\n• The system clearly shows who is present, absent, or late.\n• Shift timings are visible so staff know when to work.\n• The manager can quickly review attendance history anytime."
        ),
        (
            19,
            "Meal Time Management",
            "As a manager, I want to manage breakfast, lunch, and dinner timings so that services run smoothly without overcrowding.",
            "• Meal timings are clearly set and easy to understand.\n• Guests and staff can see when each service starts and ends.\n• The system helps avoid too many people coming at once.\n• Staff get reminders so they can prepare on time."
        ),
        (
            20,
            "Room Browsing",
            "As a customer, I want to browse available rooms so that I can choose a suitable room before booking.",
            "• Rooms are displayed in a simple and easy-to-browse way.\n• Important details like price, type, and features are clearly shown.\n• Availability is updated so users don’t see already booked rooms.\n• Images help users understand what the room looks like."
        ),
        (
            21,
            "Feedback & Reviews",
            "As a guest, I want to give feedback and reviews so that I can share my experience and help improve services.",
            "• Guests can easily leave a rating after their stay.\n• Writing a review feels simple and quick.\n• Feedback is stored so management can review it later.\n• Admin can see common issues and improve services."
        )
    ]
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    row_idx = 3
    for idx, (sn, feature, story, criteria) in enumerate(hotel_stories):
        ws_hotel.cell(row=row_idx, column=1, value=sn).alignment = Alignment(horizontal="center", vertical="top")
        ws_hotel.cell(row=row_idx, column=2, value=feature).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_hotel.cell(row=row_idx, column=3, value=story).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_hotel.cell(row=row_idx, column=4, value=criteria).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Alternating row colors
        fill_color = "F9FBFD" if idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for c in range(1, 5):
            cell = ws_hotel.cell(row=row_idx, column=c)
            cell.font = Font(name="Calibri", size=11)
            cell.fill = row_fill
            cell.border = thin_border
            
        row_idx += 1

    # Adjust columns widths for Hotel sheet
    col_widths = {1: 8, 2: 30, 3: 50, 4: 75}
    for col_num, width in col_widths.items():
        ws_hotel.column_dimensions[get_column_letter(col_num)].width = width
        
    # -------------------------------------------------------------
    # 2. Hospital Queue Management Sheet
    # -------------------------------------------------------------
    ws_hospital = wb.create_sheet(title="Hospital Queue System")
    ws_hospital.views.sheetView[0].showGridLines = True
    
    # Title Row
    ws_hospital.merge_cells('A1:D1')
    title_cell = ws_hospital['A1']
    title_cell.value = "HOSPITAL QUEUE MANAGEMENT SYSTEM — USER STORIES & ACCEPTANCE CRITERIA"
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_hospital.row_dimensions[1].height = 40
    
    # Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws_hospital.cell(row=2, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    ws_hospital.row_dimensions[2].height = 25
    
    hospital_stories = [
        (
            1,
            "Patient Registration",
            "As a Patient, I want to register and create my profile so that I can access the system and book appointments easily.",
            "• System allows patient to enter username, password, full name, contact number, age, gender, and address.\n• A unique Patient ID is generated automatically for each new patient.\n• If required fields are empty, the system shows an error message.\n• A success message is shown after registration is completed."
        ),
        (
            2,
            "Patient Login",
            "As a Patient, I want to log in securely with my credentials so that I can access my personalized role-based dashboard.",
            "• System validates the username and password combination correctly.\n• After login, the patient sees their personal dashboard only.\n• Dashboard displays patient name and relevant menu options.\n• Patient is redirected to Patient Dashboard only — not other role panels."
        ),
        (
            3,
            "Forgot Password",
            "As a User, I want to reset my password using security questions so that I can regain access to my account if I forget my password.",
            "• User can access 'Forgot Password' link from the login page.\n• System verifies identity via phone number or patient ID.\n• User must correctly answer at least 3 security questions.\n• If answers are correct, user can set a new password."
        ),
        (
            4,
            "Book Appointment",
            "As a Patient, I want to book an appointment with a doctor so that I can schedule my consultation in advance.",
            "• Patient can view list of available doctors and departments.\n• Patient can select doctor, date and preferred time.\n• System checks availability before confirming booking.\n• Booking confirmation is shown and saved in the database."
        )
    ]
    
    row_idx = 3
    for idx, (sn, feature, story, criteria) in enumerate(hospital_stories):
        ws_hospital.cell(row=row_idx, column=1, value=sn).alignment = Alignment(horizontal="center", vertical="top")
        ws_hospital.cell(row=row_idx, column=2, value=feature).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_hospital.cell(row=row_idx, column=3, value=story).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_hospital.cell(row=row_idx, column=4, value=criteria).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Alternating row colors
        fill_color = "F9FBFD" if idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for c in range(1, 5):
            cell = ws_hospital.cell(row=row_idx, column=c)
            cell.font = Font(name="Calibri", size=11)
            cell.fill = row_fill
            cell.border = thin_border
            
        row_idx += 1

    # Adjust columns widths for Hospital sheet
    for col_num, width in col_widths.items():
        ws_hospital.column_dimensions[get_column_letter(col_num)].width = width
        
    output_filename = "User_Stories_and_Acceptance_Criteria.xlsx"
    wb.save(output_filename)
    print(f"Excel file successfully generated: {os.path.abspath(output_filename)}")

if __name__ == "__main__":
    create_user_stories_excel()
